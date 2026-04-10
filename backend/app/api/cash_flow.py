from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from app.core.security import get_current_user
from app.db.connection import get_conn
import csv
import io
import openpyxl
from datetime import datetime

router = APIRouter()

def parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]

def parse_excel(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            rows.append(dict(zip(headers, row)))
    return rows

@router.post("/import")
async def import_cash_flow(
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user),
):
    content = await file.read()
    filename = file.filename.lower()
    try:
        if filename.endswith(".csv"):
            rows = parse_csv(content)
        elif filename.endswith((".xlsx", ".xls")):
            rows = parse_excel(content)
        else:
            raise HTTPException(status_code=400, detail="CSV または Excel ファイルのみ対応しています")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイルの読み込みに失敗しました: {str(e)}")

    required = {"date", "amount", "type", "description"}
    if rows and not required.issubset(set(k.lower() for k in rows[0].keys())):
        raise HTTPException(status_code=422, detail=f"必須カラムが不足しています。必要なカラム: {', '.join(required)}")

    saved = 0
    async with get_conn() as conn:
        for row in rows:
            keys = {k.lower(): v for k, v in row.items()}
            try:
                await conn.execute("""
                    INSERT INTO cash_flow_records (user_id, date, amount, type, description)
                    VALUES ($1, $2, $3, $4, $5)
                    ON CONFLICT DO NOTHING
                """,
                    user["id"],
                    datetime.strptime(str(keys["date"]), "%Y-%m-%d").date(),
                    float(keys["amount"]),
                    str(keys["type"]),
                    str(keys["description"]),
                )
                saved += 1
            except Exception:
                continue
    return {"message": f"{saved}件のデータを取り込みました", "total": len(rows), "saved": saved}

@router.get("/summary")
async def get_cash_flow_summary(
    start: str = None,
    end: str = None,
    user: dict = Depends(get_current_user),
):
    async with get_conn() as conn:
        query = """
            SELECT
                DATE_TRUNC('month', date) AS month,
                SUM(CASE WHEN type = 'income'  THEN amount ELSE 0 END) AS income,
                SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) AS expense,
                SUM(CASE WHEN type = 'income'  THEN amount
                         WHEN type = 'expense' THEN -amount ELSE 0 END) AS net
            FROM cash_flow_records
            WHERE user_id = $1
        """
        params = [user["id"]]
        if start:
            query += f" AND date >= ${len(params)+1}"
            params.append(start)
        if end:
            query += f" AND date <= ${len(params)+1}"
            params.append(end)
        query += " GROUP BY month ORDER BY month"
        rows = await conn.fetch(query, *params)
        return [
            {
                "month":   row["month"].strftime("%Y年%m月"),
                "income":  float(row["income"]),
                "expense": float(row["expense"]),
                "net":     float(row["net"]),
            }
            for row in rows
        ]
